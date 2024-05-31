from openpyxl import load_workbook

def stretch_columns_and_move_data(file_path):
    wb = load_workbook(file_path)

    # Rename the active sheet to "Dividends"
    ws = wb.active
    ws.title = "Dividends"

    # Set column widths for the "Dividends" sheet
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20

    # Create and prepare "Extracted_Data" sheet
    new_ws = wb.create_sheet(title='Extracted_Data')

    # Move data from "Dividends" to "Extracted_Data"
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=3, max_col=ws.max_column):
        new_ws.append([cell.value for cell in row])

    # Clear the data in columns C and beyond in the "Dividends" sheet
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=3, max_col=ws.max_column):
        for cell in row:
            cell.value = None

    # Set column widths for "Extracted_Data" sheet
    for col in range(1, new_ws.max_column + 1):
        col_letter = new_ws.cell(row=1, column=col).column_letter
        new_ws.column_dimensions[col_letter].width = 20

    # Save the workbook
    wb.save(file_path)
