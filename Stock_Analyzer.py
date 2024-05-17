import yfinance as yf
import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook

# Fetch and save data as before
aapl = yf.Ticker("AAPL")

# Fetch dividends data and parse 'Date' column during retrieval
dividends_df = aapl.dividends.reset_index()
# Convert 'Date' column to datetime
dividends_df['Date'] = pd.to_datetime(dividends_df['Date'])
# Convert datetime values to timezone-unaware datetimes
dividends_df['Date'] = dividends_df['Date'].dt.tz_localize(None)
# Save the data to an Excel file
dividends_df.to_excel("Apple_data_dividends.xlsx", index=False)

# Fetch quarterly cashflow data
cashflow_statement_df = aapl.quarterly_cash_flow
# Save the data to an Excel file
cashflow_statement_df.to_excel("Apple_data_quarterly_cashflow_statement.xlsx")

# Fetch quarterly balance sheet data
income_statement_df = aapl.quarterly_income_stmt
# Save the data to an Excel file
income_statement_df.to_excel("Apple_data_quarterly_income_statement.xlsx")

# Fetch quarterly balance sheet data
balance_sheet_df = aapl.quarterly_balance_sheet
# Save the data to an Excel file
balance_sheet_df.to_excel("Apple_data_quarterly_balance_sheet.xlsx")

# Add all the data into a large spreadsheet
folder_path = 'D:\\CodeProjects\\Python\\Stock-Analyzer-With-YahooFinance'
files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]

dataFrames = []
for file in files:
    df = pd.read_excel(os.path.join(folder_path, file))
    dataFrames.append(df)

merged_df = pd.concat(dataFrames)

output_path = 'D:\\CodeProjects\\Python\\Stock-Analyzer-With-YahooFinance'
merged_df.to_excel(os.path.join(output_path, 'data_combined.xlsx'), index=False)

# Stretch the columns in the combined Excel file
file_path = os.path.join(output_path, 'data_combined.xlsx')
wb = load_workbook(file_path)
ws = wb.active

# Define the width for each column
column_widths = {'A': 20, 'B': 20}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Move data from C85:H233 to a new sheet and include C1:H1 as headers
new_ws = wb.create_sheet(title='Extracted_Data')

# Copy headers (C1:H1)
headers = [cell.value for cell in ws[1][2:8]]
new_ws.append(headers)

# Copy data from C85:H233
for row in ws.iter_rows(min_row=85, max_row=233, min_col=3, max_col=8):
    new_ws.append([cell.value for cell in row])

# Clear the original data in C1:H1 and C85:H233
for cell in ws.iter_cols(min_col=3, max_col=8, min_row=1, max_row=1):
    for c in cell:
        c.value = None

for row in ws.iter_rows(min_row=85, max_row=233, min_col=3, max_col=8):
    for cell in row:
        cell.value = None

# Save the modified Excel file
wb.save(file_path)
